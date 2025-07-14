#!/usr/bin/env python3
"""Crowd management system version 48."""
from __future__ import annotations

import argparse
import asyncio
import json
import os
import threading
import time
from datetime import date, datetime
from pathlib import Path

import cv2
import torch
from fastapi import FastAPI, WebSocket, WebSocketDisconnect, Request
from fastapi.responses import StreamingResponse, HTMLResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from loguru import logger
import redis
import uvicorn
import io
import csv
from modules.tracker import FlowTracker
from modules.utils import send_email, lock, SNAP_DIR
from modules.alerts import AlertWorker

# Globals
BASE_DIR = Path(__file__).parent
TEMPLATE_DIR = BASE_DIR / "templates"

# PPE model classes
MODEL_CLASSES = [
    "dust_mask",
    "face_shield",
    "helmet",
    "protective_gloves",
    "safety_glasses",
    "safety_shoes",
    "vest_jacket",
    "head",
]

# PPE items we may track
PPE_ITEMS = [
    "helmet",
    "safety_shoes",
    "safety_glasses",
    "protective_gloves",
    "dust_mask",
    "face_shield",
    "vest_jacket",
]

ANOMALY_ITEMS = [
    "no_helmet",
    "no_safety_shoes",
    "no_safety_glasses",
    "no_protective_gloves",
    "no_dust_mask",
    "no_face_shield",
    "no_vest_jacket",
]
COUNT_GROUPS = {
    "person": ["person"],
    "vehicle": ["car", "truck", "bus", "motorcycle", "bicycle"],
}
AVAILABLE_CLASSES = (
    MODEL_CLASSES
    + ANOMALY_ITEMS
    + [c for cl in COUNT_GROUPS.values() for c in cl]
)
config_path: str
redis_client: redis.Redis
trackers: dict[int, FlowTracker] = {}
cameras: list
last_status: str | None = None
alert_worker: AlertWorker | None = None


def sync_detection_classes(cfg: dict) -> None:
    """Build class lists for object and PPE detection."""
    object_classes: list[str] = []
    count_classes: list[str] = []
    for group in cfg.get("track_objects", ["person"]):
        count_classes.extend(COUNT_GROUPS.get(group, [group]))
    object_classes.extend(count_classes)
    ppe_classes: list[str] = []
    for item in cfg.get("track_ppe", []):
        if item in MODEL_CLASSES:
            ppe_classes.append(item)
        neg = f"no_{item}"
        if neg in AVAILABLE_CLASSES:
            ppe_classes.append(neg)
    cfg["object_classes"] = object_classes
    cfg["ppe_classes"] = ppe_classes
    cfg["count_classes"] = count_classes


def load_config(path: str, r: redis.Redis) -> dict:
    if os.path.exists(path):
        data = json.load(open(path))
        data.setdefault("track_ppe", [])
        data.setdefault("alert_anomalies", [])
        data.setdefault("track_objects", ["person"])
        data.setdefault("helmet_conf_thresh", 0.5)
        data.setdefault("detect_helmet_color", False)
        data.setdefault("line_orientation", "vertical")
        data.setdefault("person_model", "yolov8n.pt")
        data.setdefault("ppe_model", "mymodalv5.pt")
        sync_detection_classes(data)
        r.set("config", json.dumps(data))
        return data
    raise FileNotFoundError(path)


def save_config(cfg: dict, path: str, r: redis.Redis) -> None:
    sync_detection_classes(cfg)
    with open(path, "w") as f:
        json.dump(cfg, f, indent=2)
    r.set("config", json.dumps(cfg))


def load_cameras(r: redis.Redis, default_url: str) -> list:
    """Load camera list from Redis or create a default one."""
    data = r.get("cameras")
    if data:
        try:
            cams = json.loads(data)
            # drop legacy class information
            for cam in cams:
                cam.pop("classes", None)
                cam.setdefault("tasks", [])
            return cams
        except json.JSONDecodeError:
            pass
    cams = [{
        "id": 1,
        "name": "Camera1",
        "url": default_url,
        "mode": "both",
        "tasks": [],
        "enabled": True,
    }]
    r.set("cameras", json.dumps(cams))
    return cams


def save_cameras(cams: list, r: redis.Redis) -> None:
    r.set("cameras", json.dumps(cams))


def reset_counts():
    """Reset all camera counts."""
    for tr in trackers.values():
        tr.in_count = 0
        tr.out_count = 0
        tr.tracks.clear()
        tr.prev_date = date.today()
        tr.redis.mset({tr.key_in: 0, tr.key_out: 0, tr.key_date: tr.prev_date.isoformat()})
    logger.info("Counts reset")

def reset_nohelmet():
    """Reset no-helmet counter."""
    redis_client.set('nohelmet_count', 0)
    logger.info("No-helmet counter reset")


def log_counts():
    ts = int(time.time())
    in_c = sum(t.in_count for t in trackers.values())
    out_c = sum(t.out_count for t in trackers.values())
    entry = json.dumps({"ts": ts, "in": in_c, "out": out_c})
    redis_client.rpush("history", entry)

async def count_log_loop():
    while True:
        log_counts()
        await asyncio.sleep(60)




def handle_status_change(status: str):
    global last_status
    if status == last_status:
        return
    last_status = status
    cfg = config.get("email", {})
    to_field = "to_yellow" if status == "yellow" else "to_red" if status == "red" else None
    if not to_field:
        return
    to_addrs = [a.strip() for a in cfg.get(to_field, "").split(',') if a.strip()]
    if not to_addrs:
        return
    threading.Thread(
        target=send_email,
        args=(f"Crowd Alert - {status.title()}", f"Current status: {status}", to_addrs, None, config.get("email", {})),
        daemon=True,
    ).start()

def start_tracker(cam: dict) -> FlowTracker:
    """Create and start a FlowTracker for a camera."""
    tr = FlowTracker(
        cam["id"],
        cam["url"],
        config.get("object_classes", ["person"]),
        config,
        cam.get("tasks", []),
        cam.get("mode", "both"),
    )
    trackers[cam["id"]] = tr
    threading.Thread(target=tr.capture_loop, daemon=True).start()
    threading.Thread(target=tr.process_loop, daemon=True).start()
    return tr


def stop_tracker(cam_id: int) -> None:
    tr = trackers.get(cam_id)
    if tr:
        tr.running = False
        trackers.pop(cam_id, None)



app = FastAPI()
templates = Jinja2Templates(directory=str(TEMPLATE_DIR))
app.mount("/snapshots", StaticFiles(directory=str(SNAP_DIR)), name="snapshots")


@app.get("/")
async def index(request: Request):
    in_c = sum(t.in_count for t in trackers.values())
    out_c = sum(t.out_count for t in trackers.values())
    current = in_c - out_c
    max_cap = config["max_capacity"]
    warn_lim = max_cap * config["warn_threshold"] / 100
    status = "green" if current < warn_lim else "yellow" if current < max_cap else "red"
    active = [c for c in cameras if c.get("enabled", True)]
    nhc = int(redis_client.get('nohelmet_count') or 0)
    return templates.TemplateResponse(
        "dashboard.html",
        {
            "request": request,
            "max_capacity": max_cap,
            "status": status,
            "current": current,
            "cameras": active,
            "cfg": config,
            "nohelmet_count": nhc,
        },
    )


@app.get("/video_feed/{cam_id}")
async def video_feed(cam_id: int):
    tr = trackers.get(cam_id)
    if not tr:
        return HTMLResponse("Not found", status_code=404)

    async def gen():
        while True:
            with lock:
                frame = tr.output_frame
            if frame is None:
                await asyncio.sleep(0.1)
                continue
            _, buf = cv2.imencode('.jpg', frame)
            yield (b'--frame\r\nContent-Type: image/jpeg\r\n\r\n' + buf.tobytes() + b'\r\n')
            await asyncio.sleep(1/tr.fps)
    return StreamingResponse(gen(), media_type='multipart/x-mixed-replace; boundary=frame')


@app.websocket('/ws/stats')
async def ws_stats(ws: WebSocket):
    await ws.accept()
    try:
        while True:
            in_c = sum(t.in_count for t in trackers.values())
            out_c = sum(t.out_count for t in trackers.values())
            current = in_c - out_c
            max_cap = config['max_capacity']
            warn_lim = max_cap * config['warn_threshold'] / 100
            status = 'green' if current < warn_lim else 'yellow' if current < max_cap else 'red'
            handle_status_change(status)
            nhc = int(redis_client.get('nohelmet_count') or 0)
            await ws.send_json({
                'in_count': in_c,
                'out_count': out_c,
                'current': current,
                'max_capacity': max_cap,
                'status': status,
                'nohelmet_count': nhc,
            })
            await asyncio.sleep(1)
    except WebSocketDisconnect:
        logger.info('WebSocket disconnected')
    except Exception as exc:
        logger.exception('Error in ws_stats: %s', exc)


@app.get('/settings')
async def settings_page(request: Request):
    return templates.TemplateResponse(
        'settings.html',
        {
            'request': request,
            'cfg': config,
            'ppe_items': PPE_ITEMS,
            'anomaly_items': ANOMALY_ITEMS,
            'count_options': list(COUNT_GROUPS.keys()),
        },
    )


@app.post('/settings')
async def update_settings(request: Request):
    data = await request.json()
    for key in [
        'max_capacity',
        'warn_threshold',
        'fps',
        'skip_frames',
        'line_ratio',
        'v_thresh',
        'debounce',
        'retry_interval',
        'conf_thresh',
        'helmet_conf_thresh',
        'detect_helmet_color',
        'person_model',
        'ppe_model',
    ]:
        if key in data:
            val = data[key]
            if key == 'detect_helmet_color':
                config[key] = bool(val) if isinstance(val, bool) else str(val).lower() == 'true'
            else:
                config[key] = type(config.get(key, val))(val)
    if 'track_ppe' in data and isinstance(data['track_ppe'], list):
        config['track_ppe'] = data['track_ppe']
    if 'alert_anomalies' in data and isinstance(data['alert_anomalies'], list):
        config['alert_anomalies'] = data['alert_anomalies']
    if 'track_objects' in data and isinstance(data['track_objects'], list):
        config['track_objects'] = data['track_objects']
    if 'line_orientation' in data:
        config['line_orientation'] = data['line_orientation']
    save_config(config, config_path, redis_client)
    for tr in trackers.values():
        tr.update_cfg(config)
    return {"saved": True}


@app.post('/reset')
async def reset_endpoint():
    reset_counts()
    return {'reset': True}


@app.post('/reset_nohelmet')
async def reset_nohelmet_endpoint():
    reset_nohelmet()
    return {'reset': True}


@app.get('/cameras')
async def cameras_page(request: Request):
    return templates.TemplateResponse(
        'cameras.html',
        {
            'request': request,
            'cams': cameras,
            'model_classes': MODEL_CLASSES,
        }
    )


@app.post('/cameras')
async def add_camera(request: Request):
    """Add a new camera without validating the stream."""
    data = await request.json()
    name = data.get('name') or f"Camera{len(cameras)+1}"
    url = data.get('url')
    mode = data.get('mode', 'both')
    tasks = data.get('tasks', [])
    if not isinstance(tasks, list):
        tasks = []
    if not url:
        return {'error': 'Missing URL'}
    cam_id = max([c['id'] for c in cameras], default=0) + 1
    cam = {
        'id': cam_id,
        'name': name,
        'url': url,
        'mode': mode,
        'tasks': tasks,
        'enabled': True,
    }
    cameras.append(cam)
    save_cameras(cameras, redis_client)
    start_tracker(cam)
    return {'added': True, 'camera': cam}


@app.delete('/cameras/{cam_id}')
async def delete_camera(cam_id: int):
    """Delete a camera by id."""
    global cameras
    remaining = [c for c in cameras if c['id'] != cam_id]
    if len(remaining) == len(cameras):
        return {'error': 'Not found'}
    cameras[:] = remaining
    stop_tracker(cam_id)
    save_cameras(cameras, redis_client)
    return {'deleted': True}


@app.patch('/cameras/{cam_id}')
async def toggle_camera(cam_id: int):
    for cam in cameras:
        if cam['id'] == cam_id:
            cam['enabled'] = not cam.get('enabled', True)
            if cam['enabled']:
                start_tracker(cam)
            else:
                stop_tracker(cam_id)
            save_cameras(cameras, redis_client)
            return {'enabled': cam['enabled']}
    return {'error': 'Not found'}


@app.put('/cameras/{cam_id}')
async def update_camera(cam_id: int, request: Request):
    data = await request.json()
    for cam in cameras:
        if cam['id'] == cam_id:
            if 'mode' in data:
                cam['mode'] = data['mode']
            if 'tasks' in data and isinstance(data['tasks'], list):
                cam['tasks'] = data['tasks']
            if 'url' in data:
                cam['url'] = data['url']
            save_cameras(cameras, redis_client)
            tr = trackers.get(cam_id)
            if tr:
                tr.update_cfg({'tasks': cam['tasks'], 'mode': cam['mode']})
            return {'updated': True}
    return {'error': 'Not found'}


@app.get('/email')
async def email_page(request: Request):
    return templates.TemplateResponse('email.html', {'request': request, 'email': config.get('email', {})})


@app.post('/email')
async def update_email(request: Request):
    data = await request.json()
    config['email'].update(data)
    save_config(config, config_path, redis_client)
    return {'saved': True}


@app.get('/email/test')
async def test_email():
    try:
        send_email('Test Email', 'This is a test email from Crowd Manager', [config['email'].get('from_addr','')], cfg=config.get('email', {}))
        return {'sent': True}
    except Exception as exc:
        return {'sent': False, 'error': str(exc)}


@app.get('/alerts')
async def alerts_page(request: Request):
    return templates.TemplateResponse('alerts.html', {
        'request': request,
        'rules': config.get('alert_rules', [])
    })


@app.post('/alerts')
async def save_alerts(request: Request):
    data = await request.json()
    config['alert_rules'] = data.get('rules', [])
    save_config(config, config_path, redis_client)
    if alert_worker:
        alert_worker.cfg = config
    return {'saved': True}


@app.get('/report')
async def report_page(request: Request):
    return templates.TemplateResponse('report.html', {'request': request})


@app.get("/report_data")
async def report_data(start: str, end: str):
    """Return logged counts for a given time range."""
    try:
        start_ts = int(datetime.fromisoformat(start).timestamp())
        end_ts = int(datetime.fromisoformat(end).timestamp())
    except Exception:
        return {"error": "invalid range"}

    entries = redis_client.lrange("history", 0, -1)
    times, ins, outs, currents = [], [], [], []
    for item in entries:
        entry = json.loads(item)
        ts = entry.get("ts")
        if ts is None or ts < start_ts or ts > end_ts:
            continue
        times.append(datetime.fromtimestamp(ts).strftime("%Y-%m-%d %H:%M"))
        i = entry.get("in", 0)
        o = entry.get("out", 0)
        ins.append(i)
        outs.append(o)
        currents.append(i - o)

    return {"times": times, "ins": ins, "outs": outs, "current": currents}

@app.get("/report/export")
async def report_export(start: str, end: str):
    data = await report_data(start, end)
    if "error" in data:
        return JSONResponse(data, status_code=400)
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Time", "In", "Out", "Current"])
    for row in zip(data["times"], data["ins"], data["outs"], data["current"]):
        writer.writerow(row)
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode()),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=report.csv"},
    )


@app.get('/ppe_report')
async def ppe_report_page(request: Request, status: str = ''):
    statuses = []
    for item in PPE_ITEMS:
        statuses.append(item)
        statuses.append(f'no_{item}')
    statuses.append('misc')
    return templates.TemplateResponse(
        'ppe_report.html',
        {
            'request': request,
            'cfg': config,
            'status': status,
            'status_options': statuses,
        },
    )


@app.get('/ppe_report_data')
async def ppe_report_data(start: str, end: str, status: str = '', min_conf: float | None = None, color: str | None = None):
    try:
        start_ts = int(datetime.fromisoformat(start).timestamp())
        end_ts = int(datetime.fromisoformat(end).timestamp())
    except Exception:
        return {"error": "invalid range"}
    entries = redis_client.lrange('ppe_logs', 0, -1)
    rows = []
    thresh = float(min_conf) if min_conf is not None else config.get('helmet_conf_thresh', 0.5)
    statuses = {s for s in status.split(',') if s}
    for item in entries:
        e = json.loads(item)
        ts = e.get('ts')
        if ts is None or ts < start_ts or ts > end_ts:
            continue
        if statuses and e.get('status') not in statuses:
            continue
        if e.get('conf', 0) < thresh:
            continue
        if color and e.get('color') != color:
            continue
        path = e.get('path')
        img_url = None
        if path:
            fname = os.path.basename(path)
            img_url = f"/snapshots/{fname}"
        rows.append({
            'time': datetime.fromtimestamp(ts).strftime('%Y-%m-%d %H:%M'),
            'cam_id': e.get('cam_id'),
            'track_id': e.get('track_id'),
            'status': e.get('status'),
            'conf': e.get('conf'),
            'color': e.get('color'),
            'image': img_url,
        })
    return {'rows': rows}


@app.get('/ppe_report/export')
async def ppe_report_export(start: str, end: str, status: str = '', min_conf: float | None = None, color: str | None = None):
    data = await ppe_report_data(start, end, status, min_conf, color)
    if 'error' in data:
        return JSONResponse(data, status_code=400)
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage
    wb = Workbook()
    ws = wb.active
    ws.append(['Time', 'Camera', 'Track', 'Status', 'Conf', 'Color', 'Image'])
    for row in data['rows']:
        ws.append([row['time'], row['cam_id'], row['track_id'], row['status'], round(row['conf'],2), row.get('color') or '' ])
        img_path = row.get('image')
        if img_path:
            img_file = os.path.join(BASE_DIR, img_path.lstrip('/'))
            if os.path.exists(img_file):
                img = XLImage(img_file)
                img.width = 80
                img.height = 60
                ws.add_image(img, f'G{ws.max_row}')
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return StreamingResponse(
        bio,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename=ppe_report.xlsx'},
    )


@app.post('/ppe_report/email')
async def ppe_report_email(start: str, end: str, status: str = '', min_conf: float | None = None, color: str | None = None, to: str | None = None):
    data = await ppe_report_data(start, end, status, min_conf, color)
    if 'error' in data:
        return JSONResponse(data, status_code=400)
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage
    wb = Workbook()
    ws = wb.active
    ws.append(['Time', 'Camera', 'Track', 'Status', 'Conf', 'Color', 'Image'])
    for row in data['rows']:
        ws.append([row['time'], row['cam_id'], row['track_id'], row['status'], round(row['conf'],2), row.get('color') or ''])
        img_path = row.get('image')
        if img_path:
            img_file = os.path.join(BASE_DIR, img_path.lstrip('/'))
            if os.path.exists(img_file):
                img = XLImage(img_file)
                img.width = 80
                img.height = 60
                ws.add_image(img, f'G{ws.max_row}')
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    recipients = [a.strip() for a in (to or config.get('email', {}).get('ppe_to', '')).split(',') if a.strip()]
    send_email(
        'PPE Report',
        'See attached report',
        recipients,
        None,
        config.get('email', {}),
        attachment=bio.getvalue(),
        attachment_name='ppe_report.xlsx',
        attachment_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    return {'sent': True}
@app.on_event('startup')
async def on_startup():
    asyncio.create_task(count_log_loop())


def main():
    global config, config_path, redis_client, cameras
    parser = argparse.ArgumentParser()
    parser.add_argument('stream_url', nargs='?')
    parser.add_argument('-c', '--config', default='config.json')
    parser.add_argument('-w', '--workers', type=int, default=None)
    args = parser.parse_args()

    config_path = args.config if os.path.isabs(args.config) else str(BASE_DIR / args.config)
    temp_cfg = json.load(open(config_path))
    redis_client = redis.Redis.from_url(temp_cfg.get('redis_url', 'redis://localhost:6379/0'))
    config = load_config(config_path, redis_client)
    cameras = load_cameras(
        redis_client,
        config['stream_url'],
    )
    if args.stream_url:
        cameras = [{
            'id': 1,
            'name': 'CameraCLI',
            'url': args.stream_url,
            'mode': 'both',
            'enabled': True,
        }]
    for cam in cameras:
        if cam.get('enabled', True):
            start_tracker(cam)

    global alert_worker
    alert_worker = AlertWorker(config, config['redis_url'], BASE_DIR)

    cores = os.cpu_count() or 1
    workers = args.workers if args.workers is not None else config['default_workers']
    w = max((cores-1 if workers == -1 else (1 if workers == 0 else workers)), 1)
    cv2.setNumThreads(w)
    torch.set_num_threads(w)
    logger.info(f"Threads={w}, cores={cores}")

    logger.info(f"Server http://0.0.0.0:{config['port']}")
    uvicorn.run(app, host='0.0.0.0', port=config['port'], log_config=None)


if __name__ == '__main__':
    main()
